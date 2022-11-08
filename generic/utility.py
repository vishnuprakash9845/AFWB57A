import openpyxl


class Excel:

    def get_cellvalue(path, sheet, row, col):
        # open the excel file
        #wb = openpyxl.load_workbook(path)
        try:
            wb = openpyxl.load_workbook(path)
        except:
            npath = path[3:]
            wb = openpyxl.load_workbook(npath)

        # goto sheet->cell->get the value
        value = wb[sheet].cell(row, col).value
        # print the value and return it
        print("XL Cell value is:", value)
        return value
